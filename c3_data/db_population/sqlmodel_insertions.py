"""
Create db population sql from exported C3 files
"""
import xml.etree.ElementTree as et
import openpyxl as px
import json
import os
import sys

# Consider the above api folder for importing sqlmodel classes
sys.path.append(os.path.join(sys.path[0], '../..'))

from sqlmodel import Session, create_engine, select, or_, and_, not_, case, col, func
import api.db as db

# Relevant filepaths
CHEMICALS_FPATH = 'c3_data/chemicals_16022021.xml'
CHEMICALS_SQL_FPATH = 'c3_data/CHEMICALS_202405021237.sql'
MONOMER_FPATH = 'c3_data/monomers.xlsx'
STOCKS_FPATH = 'c3_data/stocks.xlsx'
SCREEN_FOLDER_PATHS = ['c3_data/c3_screens', 'c3_data/commercial_screens', 'c3_data/other_screens']
PH_CURVES_FPATH = 'c3_data/manual_ph_curve_data.json'
PH_CURVES_SQL_FPATH = 'c3_data/PH_CURVE.json'
PH_POINTS_FPATH = 'c3_data/PH_POINT.json'
MISSING_MW_FPATH = 'c3_data/chemical_with_missing_MW.xlsx'

#==============================================================================#
# HELPERS
#==============================================================================#

# Convert input values (sometimes strings) to right types
def type_or_none(val, type):
    if val == '' or val == None:
        return None
    else:
        if type==int:
            return int(val)
        elif type==float:
            return float(val)
        elif type==str:
            return val
        else:
            return None

# Use new engine without logging to speed up process
engine = create_engine(db.connection_string(True), echo=False, pool_recycle=3600*2, pool_pre_ping=True)

#==============================================================================#
# API USERS
#==============================================================================#

def add_api_users():
    # Get session
    with Session(engine) as session:
        # Janet's preferred login
        user = db.ApiUser(username='janetn', password_hash=b'$2b$12$RFd3/PavfPns4OjvTvbuZeescfHgvcYBCouKpUCTgrb7oQ68mCV96', write_permission=1)
        session.add(user)
        session.commit()

#==============================================================================#
# CHEMICALS
#==============================================================================#

def insert_chemicals():
    # Get session
    with Session(engine) as session:

        # Parse file
        tree = et.parse(CHEMICALS_FPATH)
        root = tree.getroot()

        # Loop chemicals
        for chem_xml in root[0]:
            c_name = type_or_none(chem_xml.attrib['name'], str)
            c_unit = type_or_none(chem_xml.attrib['units'], str)
            c_formula = type_or_none(chem_xml.attrib['formula'], str)
            c_density = type_or_none(chem_xml.attrib['density'], float)
            c_solubility = type_or_none(chem_xml.attrib['solubility'], float)
            c_pka1 = type_or_none(chem_xml.attrib['pka1'], float)
            c_pka2 = type_or_none(chem_xml.attrib['pka2'], float)
            c_pka3 = type_or_none(chem_xml.attrib['pka3'], float)
            c_molec_weight = type_or_none(chem_xml.attrib['mw'], float)
            c_ions = type_or_none(chem_xml.attrib['ions'], str)
            c_cas = type_or_none(chem_xml.attrib['CAS'], str)
            c_cmc = type_or_none(chem_xml.attrib['CMC'], float)
            c_smiles = type_or_none(chem_xml.attrib['SMILES'], str)
            fs_concentration = type_or_none(chem_xml.attrib['def_stock_conc'], float)
            fs_unit = type_or_none(chem_xml.attrib['def_stock_units'], str)
            fs_prepip_conc = type_or_none(chem_xml.attrib['max_stock_conc'], float)
            # Add chemical
            chem = db.Chemical(name=c_name, 
                                unit=c_unit, 
                                formula=c_formula, 
                                density=c_density, 
                                solubility=c_solubility, 
                                pka1=c_pka1, 
                                pka2=c_pka2, 
                                pka3=c_pka3, 
                                molecular_weight=c_molec_weight, 
                                ions=c_ions, 
                                chemical_abstracts_db_id=c_cas, 
                                critical_micelle_concentration=c_cmc, 
                                smiles=c_smiles,
                                available=0)
            session.add(chem)
            # Get chemical id
            session.commit()
            session.refresh(chem)

            # Add frequent stock
            freq_stock = db.FrequentStock(chemical_id=chem.id, 
                                        concentration=fs_concentration, 
                                        unit=fs_unit, 
                                        precipitation_concentration=fs_prepip_conc,
                                        precipitation_concentration_unit=None if fs_prepip_conc == None else 'w/v')
            session.add(freq_stock)

            # Parse aliases and ignore classes
            for child in chem_xml:
                if child.tag == 'alias':
                    a_name = type_or_none(child.text, str)
                    # Add alias
                    alias = db.Alias(name=a_name, 
                                    chemical_id=chem.id)
                    session.add(alias)
        
        # Commit left over adds
        session.commit()

#==============================================================================#
# Stocks
#==============================================================================#

def insert_stocks():
    # Get session
    with Session(engine) as session:

        # Parse file
        wb = px.load_workbook(STOCKS_FPATH)
        l2n = lambda x: px.utils.cell.column_index_from_string(x)-1

        # Get main user id to use as creator
        creator_id = session.exec(select(db.ApiUser).where(db.ApiUser.username == 'janetn')).first()

        # Loop stocks
        for row in wb['STOCKS_300'].iter_rows(min_row=2, max_row=498):
            s_name = type_or_none(row[l2n('B')].value, str)
            f_concentration = type_or_none(row[l2n('G')].value, float)
            f_unit = type_or_none(row[l2n('H')].value, str)
            f_ph = type_or_none(row[l2n('I')].value, float)
            s_apiuser = type_or_none(creator_id, int)
            s_polar = type_or_none(1 if row[l2n('D')]=='Y' else 0, int)
            s_viscosity = type_or_none(row[l2n('J')].value, int)
            s_volatility = type_or_none(row[l2n('K')].value, int)
            s_density = type_or_none(row[l2n('U')].value, float)
            s_available = type_or_none(0 if row[l2n('M')].value==0 else 1, int)
            s_location = type_or_none(row[l2n('R')].value, str)
            s_comments = type_or_none(row[l2n('N')].value, str)
            s_hazard1 = type_or_none(row[l2n('S')].value, str)
            s_hazard2 = type_or_none(row[l2n('T')].value, str)

            # Stock file does not have a chemical name, find it from old id and sql file
            c3_chem_id = row[l2n('E')].value
            chem_found = False
            with open(CHEMICALS_SQL_FPATH, 'r') as chem_f:
                for l in chem_f.readlines():
                    l = l.strip()
                    if len(l)>1 and l.startswith('('+str(c3_chem_id)+','):
                        c_name = type_or_none(l[l.index(",'")+2:l.index("',")].strip("'"), str)
                        chem_found = True
                        break
            # Check if chemical was found
            if not chem_found:
                print('Error!: C3 Chemical ID', c3_chem_id, 'not in exported sql file')
                continue
            # Check if found chemical matches seen chemicals or chemical aliases
            chem_search = session.exec(select(db.Chemical).where(db.Chemical.name == c_name)).all()
            if len(chem_search) == 0:
                alias_search = session.exec(select(db.Alias).where(db.Alias.name == c_name)).all()
                # If can't find the chemical, ignore the stock
                if len(alias_search) == 0:
                    print('Error!: Chemical', c_name, 'not in chemical table')
                    continue
                else:
                    alias = alias_search[0]
                    chem = alias.chemical
            else:
                chem = chem_search[0]

            # Add factor if not already seen
            factor_search = session.exec(select(db.Factor).where(db.Factor.chemical_id == chem.id, 
                                                                 db.Factor.concentration == f_concentration,
                                                                 db.Factor.unit == f_unit,
                                                                 db.Factor.ph == f_ph)).all()
            if len(factor_search) == 0:
                factor = db.Factor(chemical_id=chem.id,
                                   concentration=f_concentration,
                                   unit=f_unit,
                                   ph=f_ph)
                session.add(factor)
                # Get new factor id
                session.commit()
                session.refresh(factor)
            else:
                factor = factor_search[0]
            
            # Add stock
            stock = db.Stock(factor_id=factor.id,
                             apiuser_id=s_apiuser,
                             name=s_name,
                             polar=s_polar,
                             viscosity=s_viscosity,
                             volatility=s_volatility,
                             density=s_density,
                             available=s_available,
                             location=s_location,
                             comments=s_comments)
            session.add(stock)
            # Get new stock id
            session.commit()
            session.refresh(stock)
            
            # Add hazards if not already seen or if not null
            for h_name in [s_hazard1, s_hazard2]:
                if h_name:
                    hazard_search = session.exec(select(db.Hazard).where(db.Hazard.name == h_name)).all()
                    if len(hazard_search) == 0:
                        hazard = db.Hazard(name=h_name)
                        session.add(hazard)
                        # Get new hazard id
                        session.commit()
                        session.refresh(hazard)
                    else:
                        hazard = hazard_search[0]

                    stock_hazard_link = db.Stock_Hazard_Link(stock_id=stock.id,
                                                             hazard_id=hazard.id)
                    session.add(stock_hazard_link)
            
            # Commit left over adds
            session.commit()

#==============================================================================#
# Screens
#==============================================================================#

def insert_screens():
    # Parse file
    for folder in SCREEN_FOLDER_PATHS:
        for subdir, dirs, files in os.walk(folder):

            # Make slashes consistent in case windows is being used
            subdir = subdir.replace('\\', '/')

            # Loop screens
            for file in files:
                fpath = subdir+'/'+file
                # Handle screen individually in case single screens want to be added later
                print('Processing', fpath)
                insert_screen(fpath)

def insert_screen(fpath):
    # Get session
    with Session(engine) as session:
        tree = et.parse(fpath)
        root = tree.getroot()
        if root[0].tag == 'chemicals':
            screen_xml = root[1]
        else:
            screen_xml = root[0]
        
        s_name = type_or_none(screen_xml.attrib['name'], str)
        s_available = 1
        s_owned_by = type_or_none(screen_xml.attrib['username'], str)
        s_creation_date = type_or_none(screen_xml.attrib['design_date'], str) # xml correctly formats
        s_format_name = type_or_none(screen_xml[0].attrib['name'], str)
        s_format_rows = type_or_none(screen_xml[0].attrib['rows'], int)
        s_format_cols = type_or_none(screen_xml[0].attrib['cols'], int)
        fb_reservoir_volume = type_or_none(screen_xml[0].attrib['max_res_vol'], float)
        fb_solution_volume = type_or_none(screen_xml[0].attrib['def_res_vol'], float)
        s_comments = type_or_none(screen_xml[1].text, str)

        # Add screen
        screen = db.Screen(name=s_name,
                           available=s_available,
                           owned_by=s_owned_by,
                           creation_date=s_creation_date,
                           format_name=s_format_name,
                           format_rows=s_format_rows,
                           format_cols=s_format_cols,
                           comments=s_comments)
        session.add(screen)
        # Get new screen id
        session.commit()
        session.refresh(screen)

        # Add frequent block
        if fb_reservoir_volume or fb_solution_volume:
            freq_block = db.FrequentBlock(screen_id=screen.id,
                                          reservoir_volume=fb_reservoir_volume,
                                          solution_volume=fb_solution_volume)
            session.add(freq_block)

        # Loop wellconditions
        for wellcondition_xml in screen_xml[2:]:
            wc_position_number = type_or_none(wellcondition_xml.attrib['number'], int)
            wc_label = type_or_none(wellcondition_xml.attrib['label'], str)

            # Loop factors
            f_link_dicts = []
            for factor_xml in wellcondition_xml:
                c_name = type_or_none(factor_xml.attrib['name'], str)
                f_concentration = type_or_none(factor_xml.attrib['conc'], float)
                f_unit = type_or_none(factor_xml.attrib['units'], str)
                f_ph = type_or_none(factor_xml.attrib['ph'], float)

                # Check if chemical matches matches seen chemicals or chemical aliases
                chem_search = session.exec(select(db.Chemical).where(db.Chemical.name == c_name)).all()
                if len(chem_search) == 0:
                    alias_search = session.exec(select(db.Alias).where(db.Alias.name == c_name)).all()
                    # If can't find the chemical, ignore the factor
                    if len(alias_search) == 0:
                        print('Error!: Chemical', c_name, 'not in chemical table')
                        continue
                    else:
                        alias = alias_search[0]
                        chem = alias.chemical
                else:
                    chem = chem_search[0]

                # Add factor if not already seen
                factor_search = session.exec(select(db.Factor).where(db.Factor.chemical_id == chem.id, 
                                                                    db.Factor.concentration == f_concentration,
                                                                    db.Factor.unit == f_unit,
                                                                    db.Factor.ph == f_ph)).all()
                if len(factor_search) == 0:
                    factor = db.Factor(chemical_id=chem.id,
                                    concentration=f_concentration,
                                    unit=f_unit,
                                    ph=f_ph)
                    session.add(factor)
                    # Get new factor id
                    session.commit()
                    session.refresh(factor)
                else:
                    factor = factor_search[0]
                
                # Factor links that define condition used to check if condition already exists
                f_link_dicts.append({'f_id':factor.id})

            # finding condition requires that factor link table contain all factors with the same condition id and that
            # the number of factors with that condition id being equal to the number of search factors
            clauses = []
            for f_link_dict in f_link_dicts:
                clauses.append(
                    func.sum(case((db.WellCondition_Factor_Link.factor_id == f_link_dict['f_id'], 1), else_=0)) == 1
                )
            clauses.append(func.count(db.WellCondition_Factor_Link.wellcondition_id) == len(f_link_dicts))
            # Add well condition if not already seen
            wcf_link_search = session.exec(select(db.WellCondition_Factor_Link.wellcondition_id).group_by(db.WellCondition_Factor_Link.wellcondition_id).having(*clauses)).all()
            if len(wcf_link_search) == 0:
                wellcondition = db.WellCondition(computed_similarities=0)
                session.add(wellcondition)
                # Get new well condition id
                session.commit()
                session.refresh(wellcondition)
                # Add wellcondition factor links
                for f_link_dict in f_link_dicts:
                    wcf_link = db.WellCondition_Factor_Link(wellcondition_id=wellcondition.id,
                                                            factor_id=f_link_dict['f_id'])
                    session.add(wcf_link)
                # Commit links in case condition repeats in the same screen
                session.commit()
            else:
                wellcondition = session.exec(select(db.WellCondition).where(db.WellCondition.id == wcf_link_search[0])).all()[0]

            # Add well
            well = db.Well(screen_id=screen.id,
                           wellcondition_id=wellcondition.id,
                           position_number=wc_position_number,
                           label=wc_label)
            session.add(well)
        
        # Commit left over adds
        session.commit()

#==============================================================================#
# MONOMERS
#==============================================================================#

def add_monomer_data():
    # Get session
    with Session(engine) as session:

        # Parse file
        wb = px.load_workbook(MONOMER_FPATH)
        l2n = lambda x: px.utils.cell.column_index_from_string(x)-1

        # Loop stocks
        for row in wb['Sheet1'].iter_rows(min_row=2, max_row=65):
            c_name = type_or_none(row[l2n('A')].value, str)
            c_monomer = type_or_none(row[l2n('C')].value, str)

            # Check if chemical matches matches seen chemicals or chemical aliases
            chem_search = session.exec(select(db.Chemical).where(db.Chemical.name == c_name)).all()
            if len(chem_search) == 0:
                alias_search = session.exec(select(db.Alias).where(db.Alias.name == c_name)).all()
                # If can't find the chemical, ignore the factor
                if len(alias_search) == 0:
                    print('Error!: Chemical', c_name, 'not in chemical table')
                    continue
                else:
                    alias = alias_search[0]
                    chem = alias.chemical
            else:
                chem = chem_search[0]
            
            chem.monomer = c_monomer
            session.add(chem)
        
        # Commit all updates
        session.commit()

#==============================================================================#
# PH CURVES
#==============================================================================#

def add_ph_curves():
    # Get session
    with Session(engine) as session:

        # Parse curve file
        with open(PH_CURVES_FPATH) as curves_f:
            curves = json.load(curves_f)
            for curve in curves['ph_curve']:
                pc_c_name = curve['chemical_name']
                pc_low_range = curve['low_range']
                pc_low_c_name = curve['low_chemical_name']
                pc_high_range = curve['high_range']
                pc_high_c_name = curve['high_chemical_name']
                pc_hh = curve['hh']

                # Check if chemical matches seen chemicals or chemical aliases
                chem_search = session.exec(select(db.Chemical).where(db.Chemical.name == pc_c_name)).all()
                if len(chem_search) == 0:
                    alias_search = session.exec(select(db.Alias).where(db.Alias.name == pc_c_name)).all()
                    # If can't find the chemical, ignore the factor
                    if len(alias_search) == 0:
                        print('Error!: Curve chemical', pc_c_name, 'not in chemical table')
                        continue
                    else:
                        alias = alias_search[0]
                        chem = alias.chemical
                else:
                    chem = chem_search[0]
                
                # Check if low range chemical matches seen chemicals or chemical aliases
                chem_search = session.exec(select(db.Chemical).where(db.Chemical.name == pc_low_c_name)).all()
                if len(chem_search) == 0:
                    alias_search = session.exec(select(db.Alias).where(db.Alias.name == pc_low_c_name)).all()
                    # If can't find the chemical, ignore the factor
                    if len(alias_search) == 0:
                        print('Error!: Chemical', pc_low_c_name, 'not in chemical table (curve for', pc_c_name,')')
                        continue
                    else:
                        alias = alias_search[0]
                        low_chem = alias.chemical
                else:
                    low_chem = chem_search[0]
                
                # Check if high range chemical matches seen chemicals or chemical aliases
                chem_search = session.exec(select(db.Chemical).where(db.Chemical.name == pc_high_c_name)).all()
                if len(chem_search) == 0:
                    alias_search = session.exec(select(db.Alias).where(db.Alias.name == pc_high_c_name)).all()
                    # If can't find the chemical, ignore the factor
                    if len(alias_search) == 0:
                        print('Error!: Chemical', pc_high_c_name, 'not in chemical table (curve for', pc_c_name,')')
                        continue
                    else:
                        alias = alias_search[0]
                        high_chem = alias.chemical
                else:
                    high_chem = chem_search[0]
                
                # Add ph curve
                phcurve = db.PhCurve(chemical_id=chem.id,
                                     low_range=pc_low_range,
                                     low_chemical_id=low_chem.id,
                                     high_range=pc_high_range,
                                     high_chemical_id=high_chem.id,
                                     hh=pc_hh)
                session.add(phcurve)
            
            # Commit left over adds
            session.commit()
        
        # Parse points and curve sql files
        with open(PH_CURVES_SQL_FPATH) as curves_sql_f:
            with open(PH_POINTS_FPATH) as points_f:
                curves_sql = json.load(curves_sql_f)
                points = json.load(points_f)
                seen_curves = {}
                sql_keys_with_no_curve = []
                for point in points['PH_POINT']:
                    pp_curve_sql_key = point['FK_PH_CURVE_ID']
                    pp_high_fraction = point['HIGH_PH_FRACTION_X']
                    pp_result_ph = point['RESULT_PH_Y']
                    phcurve_ids = []
                    # Points from a missing curve, skip these
                    if pp_curve_sql_key in sql_keys_with_no_curve:
                        continue
                    # Points from a seen curve
                    elif pp_curve_sql_key in seen_curves.keys():
                        phcurve_ids = seen_curves[pp_curve_sql_key]
                    # Points from as as yet unseen curve
                    else:
                        seen_curves[pp_curve_sql_key] = []
                        for curve_sql in curves_sql['PH_CURVE']:
                            if curve_sql['PK_PH_CURVE_ID'] == pp_curve_sql_key:
                                for c_name in curve_sql['NAME']:
                                    # Check if chemical matches seen chemicals or chemical aliases
                                    chem_search = session.exec(select(db.Chemical).where(db.Chemical.name == c_name)).all()
                                    if len(chem_search) == 0:
                                        alias_search = session.exec(select(db.Alias).where(db.Alias.name == c_name)).all()
                                        # If can't find the chemical, ignore the factor
                                        if len(alias_search) == 0:
                                            print('Error!: Chemical', c_name, 'not in chemical table')
                                            continue
                                        else:
                                            alias = alias_search[0]
                                            chem = alias.chemical
                                    else:
                                        chem = chem_search[0]
                                    # Note that curves are only for non-HH chemicals - which have at most one phcurve
                                    seen_curves[pp_curve_sql_key].append(chem.phcurves[0].id)
                        phcurve_ids = seen_curves[pp_curve_sql_key]
                    
                    if phcurve_ids == []:
                        print('Error!: No curves for point with SQL curve key', pp_curve_sql_key)
                        sql_keys_with_no_curve.append(pp_curve_sql_key)
                    
                    # Add point for each curve it's for
                    for phcurve_id in phcurve_ids:
                        phpoint = db.PhPoint(phcurve_id=phcurve_id,
                                             high_chemical_percentage=pp_high_fraction,
                                             result_ph=pp_result_ph)
                        session.add(phpoint)
                
                # Commit left over adds
                session.commit()

#==============================================================================#
# MOLECULAR WEIGHT
#==============================================================================#

def add_missing_mw_data():
    # Get session
    with Session(engine) as session:

        # Parse file
        wb = px.load_workbook(MISSING_MW_FPATH)
        l2n = lambda x: px.utils.cell.column_index_from_string(x)-1

        # Loop chemicals
        for row in wb['Sheet1'].iter_rows(min_row=2, max_row=102):
            c_id = type_or_none(row[l2n('A')].value, int)
            c_name = type_or_none(row[l2n('B')].value, str)
            c_mw = type_or_none(row[l2n('E')].value, float)

            # Skip when molecular weight is None
            if not c_mw:
                continue

            # Get chemical matches seen chemicals or chemical aliases
            chem = session.get(db.Chemical, c_id)
            # Double check chemical name is correct
            if chem.name != c_name:
                print('Chemical name mismatch! ID: %d, Excel: %s, DB: %s' % (c_id, c_name, chem.name))
                continue
            
            # Double check no molecular weight in db
            if chem.molecular_weight:
                print('Chemical already has a molecular weight saved! ID: %d, Name: %s' % (c_id, c_name))
                continue
            
            # Set mw
            print('Add molectular weight %lf to Chemical ID: %d, Name: %s' % (c_mw, c_id, c_name))
            chem.molecular_weight = c_mw
            session.add(chem)
        
        # Commit all updates
        session.commit()

#==============================================================================#
# Main
#==============================================================================#

# print('\nCreating Chemicals\n')
# insert_chemicals()
# print('\nCreating Stocks\n')
# insert_stocks()
# print('\nCreating Screens\n')
# insert_screens()
# print('\nAdding monomer data\n')
# add_monomer_data()
# print('\nAdding ph curves\n')
# add_ph_curves()
# print('\nAdding api users\n')
# add_api_users()
print('\nAdding missing molecular weights\n')
add_missing_mw_data()

# insert_screen('c3_data/other_screens/Design_SG2_Mol_dim.xml')